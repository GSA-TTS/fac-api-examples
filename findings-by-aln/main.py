import click
from datetime import timedelta
from openpyxl import Workbook
from peewee import *
from types import SimpleNamespace
from playhouse.shortcuts import model_to_dict
from openpyxl.styles import PatternFill
import os
from pathlib import Path

from util import (
    op,
    string_to_datetime,
    fetch_from_api,
    today,
    get_query_count
)
from const import (
    FAC_API_BASE,
    FAC_API_KEY,
    BASE_HEADERS
)

import logging
logger = logging.getLogger(__name__)

# https://stackoverflow.com/questions/17755996/how-to-make-a-list-as-the-default-value-for-a-dictionary

proxy = DatabaseProxy()  # Create a proxy for our db.

# We're going to need to cache things.
# So, a local DB makes sense.
# The table design...
# It will pull from General, Findings, and Federal Awards


class DailyGenerals(Model):
    report_id = TextField(unique=True)  # PK
    auditee_name = TextField()
    date = DateField()
    date_retrieved = DateField(null=True)
    findings_count = IntegerField(null=True)
    awards_count = IntegerField(null=True)
    cog_over = TextField(null=True)
    class Meta:
        database = proxy


class DailyFindings(Model):
    report_id = TextField()
    auditee_name = TextField()
    award_reference = TextField(null=True)
    reference_number = TextField(null=True)
    aln = TextField(null=True)
    cog_over = TextField(null=True)
    federal_program_name = TextField(null=True)
    amount_expended = IntegerField(null=True)
    is_direct = BooleanField(null=True)
    is_major = BooleanField(null=True)
    audit_report_type = TextField(null=True)
    is_passthrough_award = BooleanField(null=True)
    passthrough_amount = IntegerField(null=True)
    is_modified_opinion = BooleanField(null=True)
    is_other_matters = BooleanField(null=True)
    is_material_weakness = BooleanField(null=True)
    is_significant_deficiency = BooleanField(null=True)
    is_other_findings = BooleanField(null=True)
    is_questioned_costs = BooleanField(null=True)
    is_repeat_finding = BooleanField(null=True)
    prior_finding_ref_numbers = TextField(null=True)

    class Meta:
        database = proxy

# A result is a single award for a single day.


class QParam():
    def __init__(self, date):
        self.date = date


class Result():
    def __init__(self, d):
        self.data = SimpleNamespace(**d)

    def add(self, key, value):
        self.data[key] = value

    def __str__(self):
        return f"{self.data.report_id}"

    def __repr__(self):
        return self.__str__()


findings_fields_to_keep = set([
    "report_id",
    "award_reference",
    "reference_number",
    "is_modified_opinion",
    "is_other_matters",
    "is_material_weakness",
    "is_significant_deficiency",
    "is_other_findings",
    "is_questioned_costs",
    "is_repeat_finding",
    "prior_finding_ref_numbers",
])

awards_fields_to_keep = set([
    "federal_program_name",
    "amount_expended",
    "is_direct",
    "is_major",
    "audit_report_type",
    "is_passthrough_award",
    "passthrough_amount",
])

def convert_bools(res):
    for k in res.keys():
        if res[k] in ['Y', "TRUE", "T", "YES"]:
            res[k] = True
        elif res[k] in ["N", "NO", "FALSE", "F"]:
            res[k] = False
    return res

def get_unique_agency_numbers():
    ans = set()
    for df in DailyFindings.select():
        ans.add(df.aln.split(".")[0])
    return sorted(list(ans))

def get_unique_cog_overs():
    cogs = set()
    for df in DailyFindings.select():
        cogs.add(df.cog_over)
    return sorted(list(cogs))


def adjust_columns(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    return ws

yes_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type = "solid")

def cog_over(c, o):
    if c:
        return f"COG-{c}"
    else:
        return f"OVER-{o}"

class FAC():
    # Takes a list of parameter objects.
    def __init__(self, params):
        # TODO: Remove any dates we have already run and
        # cached locally.
        self.params = params
        self.results = []

    # Fetch the general results.
    # Must start here.
    def general(self):
        po: QParam

        for po in self.params:
            payload = {
                "fac_accepted_date": op("eq", po.date),
                "select": ",".join([
                    "report_id",
                    "auditee_name",
                    "cognizant_agency",
                    "oversight_agency"
                ]),
            }

            jres = fetch_from_api("general", payload)

            for res in jres:
                if DailyGenerals.select().where(DailyGenerals.report_id == res["report_id"]):
                    logger.debug(f"Skipping {res['report_id']}")
                else:
                    d = {"report_id": res["report_id"],
                         "date": po.date,
                         "auditee_name": res["auditee_name"],
                         "cog_over": cog_over(res["cognizant_agency"], res["oversight_agency"])
                         }
                    self.results.append(DailyGenerals.create(**d))

    # Now, populate with the findings. This tells us which we need, and
    # which to remove.
    def findings(self):
        # We should only do things where we have not fetched.
        for dg in DailyGenerals.select().where(DailyGenerals.findings_count.is_null()):
            jres = fetch_from_api("findings", {
                "report_id": op("eq", dg.report_id)
            })
            for res in jres:
                # We only need a subset of the keys
                # that come back from the API query.
                to_delete = set(res.keys()).difference(findings_fields_to_keep)
                for k in to_delete:
                    del res[k]
                # Make sure booleans are booleans...
                # Peewee does not treat 'N' as False.
                res = convert_bools(res)
                # Add the auditee name into this table.
                # Why? For human readability in the SQLite Browser.
                res["auditee_name"] = dg.auditee_name
                res["cog_over"] = dg.cog_over
                logger.debug(f"Updating with {res}")
                DailyFindings.create(**res)
            dg.date_retrieved = today()
            dg.findings_count = len(jres)
            dg.save()

    def awards(self):
        for dg in DailyGenerals.select().where(DailyGenerals.awards_count.is_null()):
            jres = fetch_from_api("federal_awards", {
                "report_id": op("eq", dg.report_id)
            })
            for res in jres:
                # Update the appropriate record.
                aln = res["federal_agency_prefix"] + \
                    "." + res["federal_award_extension"]
                award_reference = res["award_reference"]
                # We only need a subset of the keys
                # that come back from the API query.
                to_delete = set(res.keys()).difference(awards_fields_to_keep)
                for k in to_delete:
                    del res[k]
                # Make sure booleans are booleans...
                # Peewee does not treat 'N' as False.
                res = convert_bools(res)
                res["aln"] = aln
                DailyFindings.update(**res).where(DailyFindings.report_id == DailyGenerals.report_id
                                        and
                                        DailyFindings.award_reference == award_reference
                                        ).execute()

            dg.awards_count = len(jres)
            dg.save()

    def to_xlsx(self):
        wb = Workbook()
        for agency_number in get_unique_agency_numbers():
            ws = wb.create_sheet(f"{agency_number}")
            # Put headers on the sheets
            for df in (DailyFindings
                       .select
                       ().where(DailyFindings.aln.startswith(f"{agency_number}"))):
                as_d = model_to_dict(df)
                ws.append(list(as_d.keys()))
                break
            # Now the values.
            for df in (DailyFindings
                       .select
                       ().where(DailyFindings.aln.startswith(f"{agency_number}"))):
                as_d = model_to_dict(df)
                ws.append(list(as_d.values()))
            adjust_columns(ws)

        for cog_over in get_unique_cog_overs():
            ws = wb.create_sheet(cog_over)
            # Put headers on the sheets
            for df in (DailyFindings
                       .select
                       ().where(DailyFindings.cog_over == cog_over)):
                as_d = model_to_dict(df)
                ws.append(list(as_d.keys()))
                break
            # Now the values.
            for df in (DailyFindings
                       .select
                       ().where(DailyFindings.cog_over == cog_over)):
                as_d = model_to_dict(df)
                ws.append(list(as_d.values()))
            adjust_columns(ws)

        # Hyperlink the report IDs
        for sheet in wb.worksheets:
            try:
                ws = sheet
                for cell in ws["B"]:
                    if ("GSAFAC" in cell.value) or ("CENSUS" in cell.value):
                        cell.hyperlink = f"https://app.fac.gov/dissemination/report/pdf/{cell.value}"
                    else:
                        pass
                for bool_column in ["I", "J", "L", "N", "O", "P", "Q", "R", "S", "T"]:
                    for cell in ws[bool_column]:
                        if cell.value == 1:
                            cell.value = "YES"
                        elif cell.value == 0:
                            cell.value = "NO"
                for bool_column in ["J", "N", "O", "P", "Q", "R", "S", "T"]:
                    for cell in ws[bool_column]:
                        if cell.value == "YES":
                            cell.fill = yes_fill
            except:
                pass

        try:
            del wb['Sheet']
        except:
            pass
        return wb

def rm(filename):
    try:
        os.remove(filename)
    except FileNotFoundError:
        pass

@click.command()
@click.argument('acceptance_date', default="2024-03-02")
@click.option("--clean", is_flag=True, show_default=True, default=False,)
def main(acceptance_date, clean):
    acceptance_date = string_to_datetime(acceptance_date)
    db_filename = f"{acceptance_date.strftime('%Y-%m-%d')}.sqlite"
    workbook_filename = f"{acceptance_date.strftime('%Y-%m-%d')}-findings.xlsx"
    # Possibly remove work products
    if clean:
        rm(db_filename)
        rm(workbook_filename)
    # Set up the SQLite database pro
    db = SqliteDatabase(db_filename)
    proxy.initialize(db)
    db.create_tables([DailyGenerals, DailyFindings])

    qparams = []
    qparams.append(QParam(acceptance_date.date()))
    fac = FAC(qparams)

    fac.general()
    fac.findings()
    fac.awards()
    logger.info(f"Queries used: {get_query_count()}")
    wb = fac.to_xlsx()
    wb.save(workbook_filename)

if __name__ in "__main__":
    main()
