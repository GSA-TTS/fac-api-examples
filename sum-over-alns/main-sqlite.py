import os
import requests
import sys
import click
import datetime
from aln import ALN
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import sqlite3

# https://stackoverflow.com/questions/17755996/how-to-make-a-list-as-the-default-value-for-a-dictionary
from collections import defaultdict
from pprint import pprint
from collections import namedtuple as NT

FAC_API_BASE = os.getenv("FAC_API_URL")
# This change hard-overrides using the local data.
# This involves leaving out some audits, but it is faster,
# and avoids key limit issues while testing.
# FAC_API_BASE = "http://localhost:3000"
FAC_API_KEY = os.getenv("API_GOV_KEY")
MAX_RESULTS = 4_000_000
STEP_SIZE = 20000

# Basic headers; intended for use locally as well as remotely.
BASE_HEADERS = {
    "x-api-user-id": os.getenv("API_KEY_ID"),
    "authorization": f"Bearer {os.getenv('CYPRESS_API_GOV_JWT')}",
    "X-API-Key": FAC_API_KEY,
}


def load_aln_list(fname):
    alns = set()
    with open(fname, "r") as fp:
        for line in fp:
            line = line.strip()
            parts = line.split(".")
            if len(parts) == 1:
                alns.add(ALN(parts[0]))
            else:
                alns.add(ALN(parts[0], parts[1]))
    return list(alns)


def op(op, value):
    return f"{op}.{value}"


def string_to_datetime(strdate):
    parts = strdate.split("-")
    return datetime.datetime(int(parts[0]), int(parts[1]), int(parts[2]))


memoize_dates = {}


def get_date(conn, report_id):
    if memoize_dates.get(report_id, False):
        return string_to_datetime(memoize_dates.get(report_id))
    # Get the report_id and accepted date for a given report
    # payload = {
    #     "report_id": op("eq", report_id),
    #     "select": ",".join(["report_id", "fac_accepted_date"]),
    # }
    # res = requests.get(f"{FAC_API_BASE}/general", params=payload, headers=BASE_HEADERS)
    q = "SELECT report_id, fac_accepted_date from general where report_id = ?"
    r = conn.execute(q, [report_id])
    res = r.fetchone()
    if len(res) == 0:
        print(f"NO DATE FOUND FOR {report_id}")
        sys.exit()
    # ('2018-12-GSAFAC-0000011259', '2023-12-15')
    the_date = res["fac_accepted_date"]
    memoize_dates[report_id] = the_date
    the_date = string_to_datetime(the_date)
    return the_date


def calculate_for_aln(conn, aln, audit_year="2023", before_acceptance="2023-06-28"):
    # What report IDs does this ALN appear in?
    # aln : report_id
    aln_to_report_ids = defaultdict(list)
    # What is the total direct amount on that ALN?
    # aln : total
    aln_to_total = defaultdict(lambda: 0)
    # How many times do we see this ALN?
    # aln : count
    aln_to_count = defaultdict(lambda: 0)
    aln_dates = defaultdict(list)
    before_acceptance = string_to_datetime(before_acceptance)

    # We begin by finding this ALN in the federal_awards table
    # payload = {
    #     "limit": STEP_SIZE - 1,
    #     "federal_agency_prefix": op("eq", aln.agency),
    #     "audit_year": op("eq", audit_year),
    #     "is_direct": op("eq", "Y"),
    #     "select": ",".join(
    #         [
    #             "report_id",
    #             "amount_expended",
    #             "is_direct",
    #             "federal_agency_prefix",
    #             "federal_award_extension",
    #         ]
    #     ),
    # }
    # # If they included a program, and not just an agency number...
    # if aln.program:
    #     payload["federal_award_extension"] = op("eq", aln.program)

    # url = f"{FAC_API_BASE}/federal_awards"
    q = """
    SELECT 
        f.report_id, 
        f.amount_expended, 
        f.is_direct, 
        f.federal_agency_prefix, 
        f.federal_award_extension,
        g.fac_accepted_date
    FROM federal_awards f, general g
    WHERE 
        f.federal_agency_prefix = ?
        AND
        f.audit_year = ?
        AND
        f.is_direct = ?
        AND
        f.report_id = g.report_id
"""
    res = conn.execute(q, [aln.agency, audit_year, "Y"])
    all = res.fetchall()
    print(f"processing awards: {len(all)}")
    for r in all:
        # print(f"calculate_for_aln {r}")
        # this_date = get_date(conn, r["report_id"])
        # r["fac_accepted_date"] = this_date
        r["fac_accepted_date"] = string_to_datetime(r["fac_accepted_date"])
        # print(f"with date {r}")
        if r["fac_accepted_date"] < before_acceptance:
            aln_to_report_ids[aln].append(r["report_id"])
            aln_to_count[aln] = aln_to_count.get(aln, 0) + 1
            aln_dates[aln].append(r["fac_accepted_date"])
            if r["is_direct"] == "Y":
                aln_to_total[aln] = aln_to_total.get(aln, 0) + int(r["amount_expended"])

    # return (str(aln), aln_to_report_ids, aln_to_total, aln_to_count)
    return Results(
        audit_year,
        str(aln),
        aln_to_report_ids[aln],
        aln_to_total[aln],
        aln_to_count[aln],
    )


def get_alns_by_agency_number(conn, audit_year, agency_number):
    # payload = {
    #     "federal_agency_prefix": op("eq", agency_number),
    #     "select": "federal_award_extension",
    #     "audit_year": op("eq", audit_year),
    # }
    # url = f"{FAC_API_BASE}/federal_awards"
    q = "SELECT federal_award_extension from federal_awards where audit_year = ? and federal_agency_prefix = ?"
    res = conn.execute(q, [audit_year, agency_number])
    all = res.fetchall()
    all_alns = set()
    for res in all:
        print(f"alns by agency number {res}")
        all_alns.add(ALN(agency_number, res["federal_award_extension"]))
    return all_alns


def fac_weight_fun(reports, awards, dollars):
    v = (0.485 * reports) + (0.485 * awards) + (0.03 * dollars)
    return round(v, 3)


class Results:
    def __init__(self, audit_year, aln, report_ids, total_dollars, award_count):
        self.audit_year = audit_year
        self.aln = aln
        self.report_ids = set(report_ids)
        self.total_dollars = total_dollars
        self.award_count = award_count

    def __str__(self):
        return f"{self.aln} rids: {len(self.report_ids)} $: {self.total_dollars} awards: {self.award_count}"

    def __repr__(self):
        return self.__str__()

    def to_csv(self):
        return f"{self.audit_year},{self.aln},{len(self.report_ids)},{self.award_count},{self.total_dollars}"


class ResultSummary:
    def __init__(self, agency_number):
        self.agency_number = agency_number
        self.results = defaultdict(list)
        self.alns = defaultdict(list)
        self.report_counts = defaultdict(list)
        self.award_counts = defaultdict(list)
        self.total_dollars = defaultdict(list)
        self.pct_of_reports = defaultdict(list)
        self.pct_of_awards = defaultdict(list)
        self.pct_of_dollars = defaultdict(list)
        self.fac_weights = defaultdict(list)

    def add_result(self, audit_year, r):
        self.results[audit_year].append(r)

    def prep_report(self):
        for ay, rs in self.results.items():
            for r in rs:
                self.alns[ay].append(r.aln)
                self.report_counts[ay].append(len(r.report_ids))
                self.award_counts[ay].append(r.award_count)
                self.total_dollars[ay].append(int(r.total_dollars))

            self.pct_of_reports[ay] = list(
                map(
                    lambda n: round(n / sum(self.report_counts[ay]) * 100, 3),
                    self.report_counts[ay],
                )
            )
            self.pct_of_awards[ay] = list(
                map(
                    lambda n: round(n / sum(self.award_counts[ay]) * 100, 3),
                    self.award_counts[ay],
                )
            )
            self.pct_of_dollars[ay] = list(
                map(
                    lambda n: round(n / sum(self.total_dollars[ay]) * 100, 3),
                    self.total_dollars[ay],
                )
            )
            self.fac_weights[ay] = list(
                map(
                    fac_weight_fun,
                    self.pct_of_reports[ay],
                    self.pct_of_awards[ay],
                    self.pct_of_dollars[ay],
                )
            )

    def report_as_xlsx(self):
        self.prep_report()
        wb = Workbook()
        ws = wb.create_sheet("Overview")
        df = pd.DataFrame(
            {
                "note": [
                    "All values rounded to 3 places.",
                    "FAC weight is (0.485 * pct_rpt) + (0.485 * pct_awd) + (0.03 * pct_$)",
                    "FAC weight can be used for estimating opdiv contribution, if desired.",
                ]
            }
        )
        for r in dataframe_to_rows(df, index=True, header=True):
            ws.append(r)

        for ay, _ in self.results.items():
            ws = wb.create_sheet(f"AY{ay}")
            df = pd.DataFrame(
                {
                    "aln": self.alns[ay],
                    "report_count": self.report_counts[ay],
                    "award_count": self.award_counts[ay],
                    "total_dollars": self.total_dollars[ay],
                    "pct_of_reports": self.pct_of_reports[ay],
                    "pct_of_awards": self.pct_of_awards[ay],
                    "pct_of_dollars": self.pct_of_dollars[ay],
                    "fac_weight": self.fac_weights[ay],
                }
            )
            for r in dataframe_to_rows(df, index=True, header=True):
                ws.append(r)
        del wb["Sheet"]
        wb.save(f"agency-{self.agency_number}-distribution.xlsx")


def row_to_dict(cursor: sqlite3.Cursor, row: sqlite3.Row) -> dict:
    data = {}
    for idx, col in enumerate(cursor.description):
        data[col[0]] = row[idx]
    return data


@click.command()
@click.argument("list_of_alns")
@click.option("--audit-years", default="2023", help="Audit year")
@click.option("--before-acceptance", default="2023-06-28", help="Acceptance date")
@click.option(
    "--distinct-alns-for-agency",
    default=None,
    help="Each distinct aln under an agency number.",
)
def main(list_of_alns, audit_years, before_acceptance, distinct_alns_for_agency):
    RS = ResultSummary(distinct_alns_for_agency)
    conn = sqlite3.connect("fac.sqlite")
    conn.row_factory = row_to_dict

    for audit_year in list(map(lambda y: int(y), audit_years.split(","))):
        print(f"running audit year {audit_year}")
        if distinct_alns_for_agency:
            alns = get_alns_by_agency_number(conn, audit_year, distinct_alns_for_agency)
        else:
            alns = load_aln_list(list_of_alns)
        for aln in sorted(alns, key=lambda a: f"{a.agency}.{a.program}"):
            print(f"running aln: {aln}")
            result = calculate_for_aln(
                conn, aln, audit_year=audit_year, before_acceptance=before_acceptance
            )
            RS.add_result(audit_year, result)

    RS.report_as_xlsx()


if __name__ in "__main__":
    main()
